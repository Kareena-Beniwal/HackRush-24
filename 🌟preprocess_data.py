import os
import pandas as pd
from tqdm import tqdm
from glob import glob
from natsort import natsorted
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import pickle

if __name__ == '__main__':

    files = natsorted(glob('lost-in-amazon-rainforest/Train/*'))

    label_to_idx = {1:0, 2:0, 3:0, 4:1, 5:1, 6:1, 
                    7:2, 8:2, 9:2, 10:3, 11:3, 12:3,
                    13:4, 14:4, 15:4, 16:4, 17:5, 18:5,
                    19:5, 20:6, 21:6, 22:6, 23:7, 24:7, 
                    25:7, 26:8, 27:8, 28:8,}
    # Load the data
    # data = pd.read_csv('Train\data.csv')

    # # Do something with the data
    # print(data.head())
    os.makedirs('train_pkl_stride50', exist_ok=True)

    for path in files:
        complete_dataset = []
        complete_label = []
        # data = pd.read_excel(path, header=0)
        # print(data.columns)
        filename = os.path.splitext(os.path.basename(path))[0]
        sheet = load_workbook(path, read_only=True).worksheets[0]
        for i, row in enumerate(tqdm(sheet.iter_rows(values_only=True))):
            if i!=0:
                row = list(row)
                class_idx = row[-1]
                if class_idx != None:
                    curr_label = label_to_idx[class_idx]
                    complete_label.append(curr_label)
                data = row[:-1]
                complete_dataset.append(data)
    
        complete_dataset = np.array(complete_dataset)
        complete_label = np.array(complete_label)

        print(complete_dataset.shape)
        print(complete_label.shape)

        with open(f'train_pkl_stride50/{filename}_data.pkl', 'wb') as f:
            pickle.dump(complete_dataset, f)
        
        with open(f'train_pkl_stride50/{filename}_label.pkl', 'wb') as f:
            pickle.dump(complete_label, f)